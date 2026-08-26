import os
import sys

# ===== 第一时间输出友好提示，让用户知道程序正在初始化 =====
sys.stdout.write("=" * 60 + "\n")
sys.stdout.write("  宽带工单数据处理工具 v1.3\n")
sys.stdout.write("=" * 60 + "\n")
sys.stdout.write(">> 程序正在初始化，请稍候...\n")
sys.stdout.write("   (正在解压依赖库，启动可能需要 20-30 秒)\n")
sys.stdout.write("-" * 60 + "\n")
sys.stdout.flush()  # 强制立即输出

# ===== 现在开始导入标准库和第三方库 =====
import shutil
import stat
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta


def get_base_dir():
    """获取可执行文件所在目录（兼容源码运行和打包后）"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def find_files(pattern="宽带开通工单", ext=".xlsx", mode=1):
    """
    根据模式返回文件列表
    mode=1: 返回修改日期最新的一个文件（列表仅含一个）
    mode=2: 返回所有匹配的文件，按修改时间从旧到新排序
    """
    base_dir = get_base_dir()
    files = []
    for f in os.listdir(base_dir):
        if f.startswith(pattern) and f.endswith(ext):
            files.append((f, os.path.getmtime(os.path.join(base_dir, f))))
    if not files:
        return []
    files.sort(key=lambda x: x[1], reverse=True)  # 最新的在前
    if mode == 1:
        return [files[0][0]]
    else:  # mode == 2
        # 返回所有文件名，按修改时间从旧到新排序（可选）
        return [f[0] for f in files[::-1]]  # 从旧到新


def ensure_writable(filepath):
    try:
        current_mode = os.stat(filepath).st_mode
        os.chmod(filepath, current_mode | stat.S_IWUSR | stat.S_IRUSR)
    except Exception as e:
        print(f"[警告] 设置文件权限失败（非关键）：{e}")


def clean_town(value):
    if pd.isna(value):
        return value
    value_str = str(value)
    if value_str.endswith("镇"):
        return value_str[:-1]
    return value_str


def process_single_file(file_path, mapping, target_columns):
    """
    读取单个文件，过滤业务范围，映射为导用DataFrame
    返回 df_out（已处理镇区）
    """
    try:
        df_src = pd.read_excel(file_path, sheet_name="IVR未接通数据", dtype=str)
    except Exception as e:
        print(f"  [警告] 读取文件 {file_path} 失败：{e}，跳过")
        return None

    # 过滤业务范围
    allowed_business = ['FTTR开通', '宽带开通', '全屋WIFI基础产品开通']
    if '业务范围' in df_src.columns:
        df_src = df_src[df_src['业务范围'].isin(allowed_business)]
        if len(df_src) == 0:
            return None  # 无有效数据
    else:
        print(f"  [警告] 文件 {file_path} 无'业务范围'列，不过滤")

    # 映射
    df_out = pd.DataFrame(columns=target_columns)
    for src_col, tgt_col in mapping.items():
        if src_col in df_src.columns and tgt_col in target_columns:
            df_out[tgt_col] = df_src[src_col]
    # 处理镇区
    if "镇区" in df_out.columns:
        df_out["镇区"] = df_out["镇区"].apply(clean_town)
    return df_out


def main():
    # 1. 切换工作目录
    base_dir = get_base_dir()
    os.chdir(base_dir)
    print(f"\n[工作目录] {base_dir}\n")
    start_time = time.time()

    # 2. 检查模板文件
    template_file = "导数模板.xlsx"
    if not os.path.exists(template_file):
        print("[错误] 模板文件 '导数模板.xlsx' 不存在于当前目录。")
        print("   请将模板文件与 EXE 放在同一目录下。")
        input("按 Enter 键退出...")
        return

    # 3. 用户选择模式
    print("请选择处理模式：")
    print("  1 - 单文件模式（只处理最新的一份文件）")
    print("  2 - 多文件模式（合并所有匹配的文件）")
    mode_choice = input("请输入数字 1 或 2：").strip()
    while mode_choice not in ['1', '2']:
        mode_choice = input("输入无效，请重新输入 1 或 2：").strip()
    mode = int(mode_choice)

    # 4. 获取文件列表
    file_list = find_files("宽带开通工单", ".xlsx", mode)
    if not file_list:
        print("[错误] 未找到任何以 '宽带开通工单' 开头的 .xlsx 文件。")
        print(f"   请将数据文件放在当前目录：{base_dir}")
        input("按 Enter 键退出...")
        return

    if mode == 1:
        print(f"[查找] 单文件模式，处理最新文件：{file_list[0]}")
    else:
        print(f"[查找] 多文件模式，共找到 {len(file_list)} 个文件，将按日期合并处理。")

    # 5. 读取模板列名（用于映射）
    print("[模板] 正在读取模板文件列名...")
    try:
        df_template = pd.read_excel(template_file, sheet_name="导用", nrows=0)
        target_columns = df_template.columns.tolist()
        print(f"[成功] 模板列名读取成功，共 {len(target_columns)} 列\n")
    except Exception as e:
        print(f"[错误] 读取模板失败：{e}")
        input("按 Enter 键退出...")
        return

    # 6. 映射关系（所有文件共用）
    mapping = {
        "用户号码": "客户号码",
        "呼叫时间": "系统产品名称",
        "业务范围": "业务范围",
        "分公司": "分公司",
        "镇区": "镇区",
        "归档日期": "归档日期",
        "地址": "地址",
    }

    # 7. 循环处理所有文件，合并数据
    all_dfs = []
    total_files = len(file_list)
    for idx, fname in enumerate(file_list, 1):
        print(f"[处理] ({idx}/{total_files}) 正在处理文件：{fname} ...")
        df_part = process_single_file(fname, mapping, target_columns)
        if df_part is not None and len(df_part) > 0:
            all_dfs.append(df_part)
            print(f"  [成功] 该文件贡献 {len(df_part)} 条记录")
        else:
            print(f"  [跳过] 该文件无有效数据或读取失败")

    if not all_dfs:
        print("[错误] 所有文件均无有效数据，程序退出。")
        input("按 Enter 键退出...")
        return

    # 8. 合并所有数据
    df_out = pd.concat(all_dfs, ignore_index=True)
    print(f"\n[汇总] 合并后总记录数：{len(df_out)}\n")

    # 9. 生成镇区汇总（基于合并后的数据）
    if "镇区" in df_out.columns:
        print("[统计] 正在统计镇区数量...")
        town_counts = df_out["镇区"].value_counts().reset_index()
        town_counts.columns = ["镇区", "数量"]
        print(f"[成功] 共 {len(town_counts)} 个镇区\n")
    else:
        print("[错误] 合并数据中无'镇区'列，无法生成汇总。")
        input("按 Enter 键退出...")
        return

    # 10. 日期计算（汇总表使用今天的日期范围）
    today = datetime.now().date()
    start_date = today.strftime("%m%d")
    end_date = (today + timedelta(days=2)).strftime("%m%d")
    date_range = f"【{start_date}-{end_date}】新装宽带客户回访"
    today_display = f"{today.month}月{today.day}日"
    print(f"[日期] 范围：{date_range}")
    print(f"[日期] 今日：{today_display}\n")

    # 11. 构建汇总表（添加总计行）
    df_summary = pd.DataFrame({
        "日期范围": [date_range] * len(town_counts),
        "日期": [today_display] * len(town_counts),
        "镇区": town_counts["镇区"],
        "数量": town_counts["数量"]
    })
    total_count = town_counts["数量"].sum()
    summary_row = pd.DataFrame({
        "日期范围": [""],
        "日期": [""],
        "镇区": [""],
        "数量": [total_count]
    })
    df_summary = pd.concat([df_summary, summary_row], ignore_index=True)

    # ===== 生成带时间戳的输出文件名 =====
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"导数模板_处理后_{timestamp}.xlsx"
    print(f"[输出] 将生成主文件：{output_file}\n")

    # 12. 复制模板（直接复制到新文件名，不删除旧文件）
    print("[复制] 正在复制模板文件并保留样式...")
    try:
        shutil.copy2(template_file, output_file)
        print("[成功] 模板复制完成\n")
    except Exception as e:
        print(f"[错误] 复制模板失败：{e}")
        input("按 Enter 键退出...")
        return
    ensure_writable(output_file)

    # 13. 写入主文件
    print("[写入] 正在写入主文件数据...")
    try:
        wb = load_workbook(output_file)
        ws = wb["导用"]
        ws.delete_rows(2, ws.max_row)

        print("  -> 写入导用数据（全部记录）...")
        for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        ws_summary = wb.create_sheet("镇区汇总")
        print("  -> 写入镇区汇总数据...")
        for c_idx, col_name in enumerate(df_summary.columns, start=1):
            ws_summary.cell(row=1, column=c_idx, value=col_name)
        for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file)
        print(f"[成功] 主文件写入完成：{output_file}（含导用和镇区汇总）\n")
    except Exception as e:
        print(f"[错误] 写入主文件失败：{e}")
        input("按 Enter 键退出...")
        return

    # ========== 14. 按镇区分拆导出（合并所有数据） ==========
    folder_name = today.strftime("%m%d")
    try:
        if os.path.exists(folder_name):
            shutil.rmtree(folder_name)
        os.makedirs(folder_name)
        print(f"[文件夹] 创建目录：{folder_name}")
    except Exception as e:
        print(f"[错误] 创建文件夹失败：{e}")
        input("按 Enter 键退出...")
        return

    towns = df_out["镇区"].dropna().unique()
    town_list = [t for t in towns if pd.notna(t) and t != ""]
    print(f"[拆分] 开始按镇区分拆文件（合并后所有记录），共 {len(town_list)} 个镇区...\n")

    for idx, town in enumerate(town_list, 1):
        df_town = df_out[df_out["镇区"] == town].copy()
        safe_name = str(town).replace("/", "_").replace("\\", "_").replace(":", "_")
        file_path = os.path.join(folder_name, f"{safe_name}.xlsx")
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_town.to_excel(writer, sheet_name="导用", index=False)
            print(f"  [{idx}/{len(town_list)}] [成功] {town} 已保存（{len(df_town)} 条）")
        except Exception as e:
            print(f"  [{idx}/{len(town_list)}] [错误] {town} 写入失败：{e}")

    elapsed = time.time() - start_time
    print("\n" + "=" * 60)
    print("[完成] 所有任务执行完成！")
    print(f"[耗时] {elapsed:.2f} 秒")
    print(f"[主文件] {output_file}")
    print(f"[拆分目录] {folder_name}/")
    print("=" * 60 + "\n")
    input("按 Enter 键退出...")


if __name__ == "__main__":
    main()