import os
import glob
import pandas as pd
import warnings

from datetime import datetime, timedelta
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 忽略 openpyxl 的 “Workbook contains no default style” 警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ==================================================
# 配置
# ==================================================

TARGET_SHEET = "总原始表"

TARGET_COLUMNS = [
    "任务编码",
    "活动编码",
    "地市",
    "分公司",
    "网格名称",
    "网格编码",
    "责任田名称",
    "责任田编码",
    "渠道名称",
    "渠道编码",
    "活动创建时间",
    "发布人账号",
    "创建方式",
    "活动来源",
    "活动状态",
    "当前节点",
    "活动名称",
    "活动内容",
    "促销类型",
    "活动网格",
    "是否五级地址",
    "活动地点",
    "经纬度",
    "五级地址",
    "开摊时间",
    "结摊时间",
    "促销时长",
    "网格名称"
]

FIELD_MAPPING = {
    "是否五级地址": "是否关联场景库",
    "五级地址": "场景编码"
}

EMPLOYEE_FILE = "普通用户列表.xlsx"
EMPLOYEE_SHEET = "用户权限信息表"


# ==================================================
# 查找最新文件
# ==================================================

def find_latest_file(keyword, ext):
    files = glob.glob(os.path.join(os.getcwd(), f"*{keyword}*.{ext}"))
    if not files:
        raise Exception(f"未找到文件：{keyword}*.{ext}")
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


# ==================================================
# 获取日期
# ==================================================

def get_target_date():
    value = input("请输入日期(yyyymmdd，例如20260729)，直接回车默认昨天：").strip()
    if value:
        try:
            datetime.strptime(value, "%Y%m%d")
            return value
        except:
            raise Exception("日期格式错误")
    yesterday = datetime.now() - timedelta(days=1)
    result = yesterday.strftime("%Y%m%d")
    print(f"未输入日期，默认使用昨天：{result}")
    return result


# ==================================================
# 判断日期
# ==================================================

def check_date(value, target):
    if pd.isna(value):
        return False
    value = str(value)
    value = value.replace("-", "").replace("/", "").replace(" ", "").replace(":", "")
    return target in value


# ==================================================
# 复制Excel格式
# ==================================================

def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


# ==================================================
# 加载员工 ID 映射（Portal → 工号）
# ==================================================

def load_employee_mapping():
    try:
        df_emp = pd.read_excel(EMPLOYEE_FILE, sheet_name=EMPLOYEE_SHEET, dtype=str)
        df_emp.columns = df_emp.columns.astype(str).str.strip()
        mapping = {}
        for _, row in df_emp.iterrows():
            portal = str(row.get('Portal', '')).strip()
            emp_id = str(row.get('工号', '')).strip()
            if portal and emp_id:
                mapping[portal] = emp_id
        return mapping
    except Exception as e:
        print(f"警告：读取员工表失败，将不进行 portal 转换。错误：{e}")
        return {}


# ==================================================
# 生成工号次数统计文件
# ==================================================

def generate_attendance_stats(result_df, target_date, portal_to_emp):
    if result_df.empty:
        print("没有数据，跳过统计文件生成")
        return

    required_cols = ['任务编码', '开摊时间', '经纬度', '网格名称', '分公司']
    for col in required_cols:
        if col not in result_df.columns:
            print(f"警告：result_df 缺少列 '{col}'，统计文件可能不完整")

    src_resp_col = None
    src_part_col = None
    for col in result_df.columns:
        if '活动负责人' in col:
            src_resp_col = col
            break
    for col in result_df.columns:
        if '活动参与人' in col:
            src_part_col = col
            break

    rows = []
    for _, row in result_df.iterrows():
        task_code = row.get('任务编码', '')
        open_time = row.get('开摊时间', '')
        lnglat = row.get('经纬度', '')
        grid_name = row.get('网格名称', '')
        branch = row.get('分公司', '')

        try:
            dt = pd.to_datetime(open_time)
            open_date = dt.strftime('%Y%m%d')
        except:
            open_date = str(open_time).replace('-', '').replace('/', '').replace(' ', '').replace(':', '')[:8]

        emp_ids = []
        if src_resp_col and pd.notna(row.get(src_resp_col)):
            val = str(row[src_resp_col]).strip()
            if val:
                emp_ids.append(val)
        if src_part_col and pd.notna(row.get(src_part_col)):
            val = str(row[src_part_col]).strip()
            if val:
                parts = [p.strip() for p in val.replace('，', ',').split(',') if p.strip()]
                emp_ids.extend(parts)

        for emp in emp_ids:
            emp_final = portal_to_emp.get(emp, emp)
            rows.append({
                '任务编码': task_code,
                '工号': emp_final,
                '开摊时间': open_date,
                '工号+时间': f"{emp_final}{open_date}",
                '经纬度': lnglat,
                '网格': grid_name,
                '分公司': branch
            })

    if not rows:
        print("没有提取到任何工号，统计文件为空")
        return

    df_stats = pd.DataFrame(rows)
    counts = df_stats.groupby('工号').size().reset_index(name='次数')
    df_stats = df_stats.merge(counts, on='工号', how='left')

    cols_order = ['任务编码', '工号', '开摊时间', '工号+时间', '经纬度', '次数', '网格', '分公司']
    for col in cols_order:
        if col not in df_stats.columns:
            df_stats[col] = ''
    df_stats = df_stats[cols_order]

    dt = datetime.strptime(target_date, '%Y%m%d')
    month_day = f"{dt.month}.{dt.day:02d}"
    filename = f"工号次数{month_day}.xlsx"
    df_stats.to_excel(filename, index=False)
    print(f"已生成统计文件：{filename}")


# ==================================================
# 解析表头，支持重复列名（返回列名→列号列表）
# ==================================================

def parse_headers(ws):
    """
    解析 Excel 第一行，返回一个字典：
    {列名: [列号列表]}
    支持重复列名。
    """
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col_idx).value
        if cell_value is not None:
            col_name = str(cell_value).strip()
            if col_name not in header_map:
                header_map[col_name] = []
            header_map[col_name].append(col_idx)
    return header_map


# ==================================================
# 主流程
# ==================================================

def process():
    target_date = get_target_date()

    portal_to_emp = load_employee_mapping()
    def convert_to_emp_id(value):
        if pd.isna(value):
            return value
        s = str(value).strip()
        return portal_to_emp.get(s, s)

    csv_file = find_latest_file("现场摆摊记录", "csv")
    excel_file = find_latest_file("促销打卡", "xlsx")

    print("==========正在读取文件============")
    print("CSV文件：")
    print(csv_file)
    print()
    print("Excel文件：")
    print(excel_file)
    print("================================")

    # 读取CSV
    df = pd.read_csv(csv_file, dtype=str, encoding="utf-8-sig")
    df.columns = df.columns.astype(str).str.strip()

    src_resp_col = None
    src_part_col = None
    for col in df.columns:
        if '活动负责人' in col:
            src_resp_col = col
            break
    for col in df.columns:
        if '活动参与人' in col:
            src_part_col = col
            break

    required = ["活动状态", "开摊时间", "结摊时间", "活动来源"]
    for col in required:
        if col not in df.columns:
            raise Exception(f"CSV缺少字段：{col}")

    result = df[
        (df["活动状态"] == "已完成") &
        (df["开摊时间"].apply(lambda x: check_date(x, target_date))) &
        (df["结摊时间"].apply(lambda x: check_date(x, target_date))) &
        (df["活动来源"] == "网格调度工作台")
    ].copy()

    print(f"筛选符合条件数据：{len(result)}条")
    if len(result) == 0:
        print("没有符合条件数据，将退出")
        return

    # 生成统计文件
    generate_attendance_stats(result, target_date, portal_to_emp)

    # ==================== 新增进度提示：准备打开 Excel ====================
    print("正在打开 Excel 文件并准备写入数据...")
    # ====================================================================

    # 打开Excel并写入
    wb = load_workbook(excel_file)
    if TARGET_SHEET not in wb.sheetnames:
        raise Exception(f"不存在Sheet：{TARGET_SHEET}")
    ws = wb[TARGET_SHEET]

    # 使用新的表头解析函数（支持重复列名）
    header_map = parse_headers(ws)

    # 获取所有工号列（按数字排序）
    worker_cols = []
    for col_name in header_map.keys():
        if col_name.startswith('工号') and col_name[2:].isdigit():
            worker_cols.append(col_name)
    worker_cols.sort(key=lambda x: int(x[2:]))

    last_row = ws.max_row
    while last_row > 1:
        values = [ws.cell(last_row, c).value for c in range(1, ws.max_column + 1)]
        if any(values):
            break
        last_row -= 1

    write_row = last_row + 1
    template_row = last_row

    # ==================== 新增进度提示：开始写入循环 ====================
    total_rows = len(result)
    print(f"正在写入 {total_rows} 条数据到工作表 '{TARGET_SHEET}'，请稍候...")
    # ====================================================================

    for idx, (_, row) in enumerate(result.iterrows(), 1):
        copy_row_format(ws, template_row, write_row)

        # 写入目标列（支持重复列名）
        for col in TARGET_COLUMNS:
            if col not in header_map:
                continue

            value = ""
            if col == "网格名称":
                if "网格名称" in result.columns:
                    value = row["网格名称"]
            elif col == "促销时长":
                start_col = header_map["开摊时间"][0]
                end_col = header_map["结摊时间"][0]
                value = f"={get_column_letter(end_col)}{write_row}-{get_column_letter(start_col)}{write_row}"
            elif col in FIELD_MAPPING:
                source = FIELD_MAPPING[col]
                if source in result.columns:
                    value = row[source]
            elif col in result.columns:
                raw_val = row[col]
                if col in ("开摊时间", "结摊时间", "活动创建时间") and pd.notna(raw_val):
                    try:
                        dt = pd.to_datetime(raw_val)
                        if pd.notna(dt):
                            value = dt.to_pydatetime()
                        else:
                            value = raw_val
                    except Exception:
                        value = raw_val
                else:
                    value = raw_val

            for col_idx in header_map[col]:
                ws.cell(write_row, col_idx).value = value

        # 工号列写入
        if src_resp_col and len(worker_cols) >= 1:
            resp_val = row.get(src_resp_col, '')
            if pd.notna(resp_val) and str(resp_val).strip():
                col_name = worker_cols[0]
                for col_idx in header_map.get(col_name, []):
                    ws.cell(write_row, col_idx).value = convert_to_emp_id(resp_val)

        if src_part_col and len(worker_cols) >= 2:
            part_val = row.get(src_part_col, '')
            if pd.notna(part_val):
                parts = [p.strip() for p in str(part_val).replace('，', ',').split(',') if p.strip()]
                for i, part in enumerate(parts):
                    if i + 1 < len(worker_cols):
                        col_name = worker_cols[i+1]
                        for col_idx in header_map.get(col_name, []):
                            ws.cell(write_row, col_idx).value = convert_to_emp_id(part)

        write_row += 1

        # ==================== 新增进度提示：每10条或最后一条打印进度 ====================
        if idx % 10 == 0 or idx == total_rows:
            print(f"  进度：{idx}/{total_rows} 条")
        # =================================================================================

    # ==================== 新增进度提示：写入完成，准备保存 ====================
    print("数据写入完成，正在保存文件...")
    # =========================================================================

    # 强制右对齐
    align_right_columns = ["五级地址", "开摊时间", "结摊时间"]
    for col_name in align_right_columns:
        if col_name in header_map:
            for col_idx in header_map[col_name]:
                for row_idx in range(last_row + 1, write_row):
                    cell = ws.cell(row_idx, col_idx)
                    cell.alignment = Alignment(horizontal='right')

    wb.save(excel_file)

    # ==================== 新增进度提示：保存成功 ====================
    print("文件保存成功！")
    # ================================================================

    print("======================")
    print("处理完成")
    print(f"新增数据：{len(result)}条")
    print(f"已更新文件：{excel_file}")
    print("======================")


if __name__ == "__main__":
    process()